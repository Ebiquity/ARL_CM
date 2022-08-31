""" Convert Bridging combat messages spreadsheet to an RDF knowledge graph """

import argparse
from os.path import exists
from openpyxl import load_workbook, utils
from rdflib import Graph, URIRef, Namespace, BNode, Literal
from rdflib.namespace import XSD, RDF, OWL, RDFS
import re
import mgrs
import shortuuid 
from collections import defaultdict

## global variables
GRAPH = None        # the rdf graph
SIMULATION = None   # the id of the simulation in the graph
TASK = None         # the id of the task in the graph

# keeps track of messages we don't understand
unrecognized_alerts = defaultdict(int)

# namespaces
cm = "http://purl.org/artiamas/cm/"
CM = Namespace(cm)

# dict from objects seen in Target field to their IRIs
objname2IRI = {}

# link actions to their corresponding messages. If not linked, we can find
# a corresponding one using the sequence proporty.
link_msgs_actions = False

def warn(msg):
    # added to distinguish from debugging or status print statements
    print(msg)

def bnode(prefix = ''):
    # custom BBode function adds a prefix to a short uuid sequence
    if prefix:
        return BNode(prefix + '_' + shortuuid.uuid()[:5])
    else:
        return BNode(shortuuid.uuid()[:5])

def getObj(label):
    """ returns IRI for task or object found in the Target field, creating it if neccessary """
    if label in objname2IRI:
        return objname2IRI[label]
    # not seen before, so create it
    if "Crossing Task" in label:
        obj = bnode('TASK')
        GRAPH.add((obj, RDF.type, CM.CrossingTask))
        GRAPH.add((SIMULATION, CM.task, obj))
    elif "River Terrain" in label:
        obj = bnode('OBJ')
        GRAPH.add((obj, RDF.type, CM.Obstacle))
    else:
        warn('Unrecognized object:', label)
        objname2IRI[label] = None
        return None
    GRAPH.add((obj, RDFS.label, Literal(label)))
    objname2IRI[label] = obj
    return obj

agent_instances = {'':CM.NONE} # dict of names to bnodes

#file = load_workbook(filename = cm_sheet)
#sheet = file.active

def sheet2json(sheet_file):
    """ read the spreadsheet data into a JSON-like python structure,
    i.e. a list of dictionaries, one for eeach data row.  Each
    dictionary has keys corresponding to the column header values
    corresponding to the cell value """
    sheet = load_workbook(sheet_file).active
    allData = []
    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))
    for row in range(2, last_row + 1):
        data = {}
        data['sequence'] = row-1
        for column in range(1, last_column + 1):
            col = utils.get_column_letter(column)
            property = sheet[col + str(1)].value
            value = sheet[col + str(row)].value
            if row > 1:
                data[property] = value
        allData.append(data)
    return allData

# dictionary mapping a spreadsheet column name to a RDF property URI
str2property = {'Time':CM.time, 'Agent':CM.agent, 'sequence':CM.sequence, 
                'Type':CM.messageType, 'Level':CM.level, 'Grid':CM.grid,
                'Agent':CM.agent, 'Target':CM.target, 'Alert Messages':CM.alertMessage,
                'Force':CM.force, 'id':CM.id}

    
def str2object(msg, prop, value):
    """ convert a string into a object or literal so we can map a
        message's prop value to a RDF property URI """
    if prop == 'Force':
        return force_value(value)
    elif type(value) in [int, float]:
        return Literal(value)
    elif prop == 'Agent':
        return parse_agent(value, msg=msg)
    elif prop == 'Target':
        if value and "[1]" in value:
            # sometimes the Target field has an object reference (e.g., task or obstacle)
            return getObj(value)
        elif not msg['Agent']:
            return parse_agent(value, msg=msg)
        else:
            return parse_target(value, msg=msg)
    else:
        return Literal(value)
    
def force_value(value, opposite=False):
    # maps a spreadsheet force value to the cm: force value
    if value in ["B", "BLUFOR"]:
        return Literal("red") if opposite else Literal("blue")
    elif value in ["R", "REDFOR"]:
        return Literal("blue") if opposite else Literal("red")
    else:
        warn('Unrecognized force value:', value)
        return Literal("unknown")

### Given a string (e.g., "B CO / 1 - 22") representing a military
### unit we've not seen before, we parse the string to infer the
### unit's type (e.g., CM.Company) and its superior units (e.g.,
### Battalion 1; Regiment 22) nd their types, adding information to
### the RDF graph. The function returns the sgent's URI

def normalize_unit_name(text):
    """ returns a name with a slash between components,
     e.g., SCT_PLT/1/22_IN """
    if not text:
        return ''
    name = text.strip()
    if not name:
        return ''
    if '/' not in name:
        # sometimes there's just the local name
        name = name + ' / 1 / 22 IN'
    name = name.replace('-','/')
    name = name.replace(' ','_')
    if re.search('^\d_\d_CO', name):
        name = name[0] + '/' + name[1:]
    unit = [x.strip('_') for x in name.split('/')]
    name = '/'.join(unit)
    return name

def infer_unit_type_and_parent(name):
    """ returns unit type based on its name, e.g.; SCT_PLT/1/22_IN  => CM.ScoutPlatoon 
    standard hierarchy: squad<platoon<company<battalion<brigade|regiment<division<corps
    """
    units = name.split('/')
    unit0 = units[0]
    parent = '/'.join(units[1:])
    if 'ENG_CO' in unit0:
        unit_type = CM.EngineeringCompany
    elif 'SCT_PLT' in unit0:
        unit_type = CM.ScoutPlatoon
    elif 'MORTAR_PLT' in unit0:
        unit_type = CM.MortarPlatoon
    elif 'CO' in unit0:
        unit_type = CM.Company
    elif len(units) > 1 and 'CO' in units[1]:
        unit_type = CM.Platoon
    else:
        warn('Unrecognized unit type:', name)
        unit_type = CM.MilitaryUnit
    return (unit_type, parent)

def parse_target(text, field='target', infer_units=True, msg=None):
    return parse_agent(text, field='target', infer_units=infer_units, msg=msg)

def parse_agent(text, field='agent', infer_units=True, msg=None):
    """returns entity iri id """
    global agent_instances

    name = normalize_unit_name(text)
    if name in agent_instances:
        # We've seen this before, so just return the instance
        return agent_instances[name]
    
    unit_type, parent = infer_unit_type_and_parent(name)
      
    id = URIRef(cm + 'UNIT_'+ name)
    agent_instances[name] = id
    if field == 'agent':
        # agent's force is same as msg force
        force = force_value(msg['Force'])
        GRAPH.add((id, CM.force, force))
    elif field == 'target' and msg and msg['Agent']:
        # target's force (if there is an agent) opposite of msg force
        force = force_value(msg['Force'], opposite=True)
        GRAPH.add((id, CM.force, force ))
    GRAPH.add((id, CM.isa, unit_type))
    GRAPH.add((id, RDFS.label, Literal(name)))
    
    if infer_units:
        infer_superunits(id, unit_type, force, parent)
    return id
  
def get_unit_type(t):
    if 'Platoon' in t:
        return 'platoon'
    elif 'Company' in t:
        return 'company'
    elif 'Brigade' in t:
        return 'brigade'
    elif 'Battalion' in t:
        return 'battalion'
    elif 'Regiment' in t:
        return 'regiment'
    else:
        warn('Unrecognized unit type:', t)
        return None

def infer_superunits(subid, subunit_type, force, name):
    if not name:
        return None
    id = URIRef(cm + 'UNIT_'+ name)
    agent_instances[name] = id
    GRAPH.add((id, CM.force, force))
    GRAPH.add((id, RDFS.label, Literal(name)))
    GRAPH.add((subid, CM.partOf, id))
    
    units = name.split('/')
    unit0 = units[0]
    parent = '/'.join(units[1:])
    
    subtype = get_unit_type(subunit_type)
    
    if subtype == "platoon":
        # platoon can be part of a company or battalion
        if 'CO' in unit0:
            unit_type = CM.Company
        else:
            unit_type = CM.Battalion
        GRAPH.add((id, CM.isa, unit_type))
    elif subtype == "company":
        # company always part of a battalion
        unit_type = CM.Battalion
        GRAPH.add((id, CM.isa, unit_type))
    elif subtype == "battalion":
        # battalion can be part of a brigade or regiment
        if "INF BDE" in unit0:
            unit_type = CM.InfantryBrigade
        elif "BDE" in unit0:
            unit_type = CM.Brigade
        elif "IN" in unit0:
            unit_type = CM.InfantryRegiment
        else:
            unit_type = CM.Regiment
        GRAPH.add((id, CM.isa, unit_type))
    else:
        warn(f"Unrecognized subunit type {unit0} for ({subunit_type}, {force}, {name}")
        return None
    if parent:
        infer_superunits(id, unit_type, force, parent)

def duration(time1, time2):
    # won't work if we cross a day boundry :-( 
    # H+08:12 
    h1,m1  = time1.split('+')[1].split(':')
    h2,m2  = time2.split('+')[1].split(':')
    return str(h2-h1) + ':' + str(m2-m1)
    
def time_to_minutes(time):
    h, m  = [int(x) for x in time.split('+')[1].split(':')]
    return 60*h + m


def add_geopoint(grid, m = mgrs.MGRS()):
    # converts a grid reference to a lat/long tuple and returns a reference
    # a geopoint containing them
    lat, lon = m.toLatLon(grid.replace(' ',''))  
    gp = bnode('GEO')
    GRAPH.add((gp, CM.grid, Literal(grid)))
    GRAPH.add((gp, CM.latitude, Literal(lat, datatype=XSD.double)))
    GRAPH.add((gp, CM.longitude, Literal(lon, datatype=XSD.double)))
    return gp

def add_action(msg, msg_id):
    """ add an action to the graph based on the message"""
    act = bnode('ACT')
    alert = msg["Alert Messages"].lower()
    force = msg["Force"]
    agent = list(GRAPH.objects(msg_id, CM.agent))[0]
    target = list(GRAPH.objects(msg_id, CM.target))[0]
    time = list(GRAPH.objects(msg_id, CM.time))[0]
    minutes = list(GRAPH.objects(msg_id, CM.minutes))[0]
    geopoint = add_geopoint(list(GRAPH.objects(msg_id, CM.grid))[0])
    if link_msgs_actions:
        GRAPH.add((act, CM.message, msg_id))
    GRAPH.add((act, CM.time, list(GRAPH.objects(msg_id, CM.time))[0] ))
    GRAPH.add((act, CM.minutes, list(GRAPH.objects(msg_id, CM.minutes))[0] ))
    GRAPH.add((act, CM.sequence, list(GRAPH.objects(msg_id, CM.sequence))[0] ))
    GRAPH.add((act, CM.alertMessage, Literal(alert)))
    
    if 'resupply' in alert:
        # a unit is resupplied with ammo, fues or something else
        GRAPH.add((act, CM.isa, CM.Resupply))
        supplier = CM.BlueSupplyUnit if force == "B" else CM.RedSupplyUnit
        GRAPH.add((act, CM.force, Literal(force)))
        GRAPH.add((act, CM.subject, supplier))
        GRAPH.add((act, CM.recipient, target))
        if "(ammo)" in alert:
            GRAPH.add((act, CM.object, CM.Ammo))
        elif "(fuel)" in alert:
            GRAPH.add((act, CM.object, CM.Fuel))
        else:
            GRAPH.add((act, CM.object, CM.Supplies))
    elif 'cbe' in alert:
        # CBE = close battle event
        if 'movement resumed' in alert:
            GRAPH.add((act, CM.isa, CM.Move))
            GRAPH.add((act, CM.subject, agent))
            GRAPH.add((act, CM.toward, target))
            act1 = bnode('ACT')
            GRAPH.add((act, CM.reason, act1))
            GRAPH.add((act, CM.object, act1 ))
            GRAPH.add((act1, CM.isa, CM.CBE))
            GRAPH.add((act1, CM.subject, agent))
            GRAPH.add((act1, CM.object, target))
        elif 'unit paused' in alert:
            GRAPH.add((act, CM.isa, CM.Pause))
            GRAPH.add((act, CM.subject, agent))
            GRAPH.add((act, CM.reason, Literal(alert)))
            act1 = bnode('ACT')
            GRAPH.add((act, CM.object, act1 ))
            GRAPH.add((act1, CM.isa, CM.CBE))
            GRAPH.add((act1, CM.subject, agent))
            GRAPH.add((act1, CM.object, target))
        elif 'terminated' in alert:
            GRAPH.add((act, CM.isa, CM.Cancel))
            GRAPH.add((act, CM.subject, agent))
            GRAPH.add((act, CM.reason, Literal(alert)))
            act1 = bnode('ACT')
            GRAPH.add((act, CM.object, act1 ))
            GRAPH.add((act1, CM.isa, CM.CBE))
            GRAPH.add((act1, CM.subject, agent))
            GRAPH.add((act1, CM.object, target))
        else:
            warn(f"Unrecognized CBE alert: {alert}")
    elif ' sa ' in alert:
        # unit gaining or losing situational awareness of an enemy unit
        actType = CM.EarnSA if 'earned' in alert else CM.LostSA
        GRAPH.add((act, CM.isa, actType))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
    elif 'moving into range' in alert:
        reason = Literal(alert[15:].strip(', '))
        GRAPH.add((act, CM.isa, CM.Move))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.reason, reason))
    elif 'in df range' in alert:
        actType = CM.Attack if '(attacking)' in alert else CM.Engage
        GRAPH.add((act, CM.isa, actType))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
    elif 'moving to fight' in alert or          'adjust route to fight' in alert:
        GRAPH.add((act, CM.isa, CM.Move))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.reason, Literal("attack")))
    elif 'firing (artillery)' in alert:
        GRAPH.add((act, CM.isa, CM.Firing))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        if "on" in alert:
            GRAPH.add((act, CM.startTime, time))
        elif "ended" in alert:
            GRAPH.add((act, CM.endTime, time))
        else:
            warn(f"Unrecognized action {alert}")
    elif re.match('attacking.*against', alert):
        GRAPH.add((act, CM.isa, CM.Attack))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
    elif 'fighting' in alert:
        GRAPH.add((act, CM.isa, CM.StartFight))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.toward, target))
    elif re.match('attacking.*ended', alert):
        GRAPH.add((act, CM.isa, CM.EndFight))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.toward, target))
    elif re.match('receiving.*fire$', alert):
        GRAPH.add((act, CM.isa, CM.Attack))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.startTime, time))
    elif re.match('receiving.*fire ended', alert):
        GRAPH.add((act, CM.isa, CM.Attack))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.endTime, time))
    elif "not going after opfor" in alert:
        # OPFOR is "opposing force"
        GRAPH.add((act, CM.isa, CM.Stop))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.status, Literal("unable")))
        GRAPH.add((act, CM.reason, Literal(alert)))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Engage))
        GRAPH.add((act1, CM.subject, agent))
        GRAPH.add((act1, CM.object, target))
    elif "can't create directfires" in alert:
        # fixme
        GRAPH.add((act, CM.isa, CM.Report))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.status, Literal(alert)))
    elif "paused at crossing control point" in alert:
        GRAPH.add((act, CM.isa, CM.Pause))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.location, geopoint))
        #GRAPH.add((act, CM.status, Literal("pause" )))
        if "crossing is not traversable" in alert:
            GRAPH.add((act, CM.reason, Literal("crossing not traversable")))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Move))
        GRAPH.add((act1, RDFS.label, Literal("a move")))
        GRAPH.add((act1, CM.subject, agent))
        GRAPH.add((act1, CM.object, target))
    elif 'crossing begin' in alert:
        # a unit is starting to cross
        GRAPH.add((act, CM.isa, CM.Crossing))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.startTime, time))
        GRAPH.add((act, CM.minutes, minutes))
    elif 'crossing complete' in alert:
        # a unit has completed a crossing
        # find corresponding act?
        GRAPH.add((act, CM.isa, CM.Crossing))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        GRAPH.add((act, CM.endTime, time))
        GRAPH.add((act, CM.minutes, minutes))
    elif 'task' in alert or 'crossing' in alert:
        # messages about the Crossing Task
        #GRAPH.add((act, CM.isa, CM.CrossingTaskAction))
        #GRAPH.add((act, CM.subject, agent))
        #GRAPH.add((act, CM.object, TASK))
        # note, here act1 is the top-level act, not act
        act1 = bnode('ACT')
        GRAPH.add((act, CM.isa, CM.CrossingTaskAction))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.location, geopoint))
        GRAPH.add((act, CM.object, TASK))
        if 'start' in alert or 'begin' in alert:
            GRAPH.add((act1, CM.isa, CM.Begin))
            GRAPH.add((act1, CM.time, time))
            GRAPH.add((act1, CM.minutes, minutes))
            GRAPH.add((act1, CM.agent, agent))
            GRAPH.add((act1, CM.object, act))
            GRAPH.add((act1, CM.location, geopoint))
        elif 'complete' in alert:
            GRAPH.add((act1, CM.isa, CM.Finish))
            GRAPH.add((act1, CM.time, time))
            GRAPH.add((act1, CM.minutes, minutes))
            GRAPH.add((act1, CM.agent, agent))
            GRAPH.add((act1, CM.object, act))
            GRAPH.add((act, CM.status, Literal("ended")))
            GRAPH.add((TASK, CM.status, Literal("completed")))
            GRAPH.add((TASK, CM.endTime, time))
    elif 'pause' in alert and 'obstacle' in alert:
        GRAPH.add((act, CM.isa, CM.Pause))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.object, target))
        #GRAPH.add((act, CM.status, Literal("pause")))
        GRAPH.add((act, CM.reason, target))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Move))
        GRAPH.add((act1, CM.subject, agent))
    elif 'waiting' in alert:
        GRAPH.add((act, CM.isa, CM.Pause))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.status, Literal("wait" )))
        if "can't attack without reinforcements" in alert:
            GRAPH.add((act, CM.reason, Literal("need reinforcements")))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.oject, act1 ))
        GRAPH.add((act1, CM.isa, CM.Attack))
        GRAPH.add((act1, CM.subject, agent))
        GRAPH.add((act1, CM.object, target))
    elif 'resume' in alert:
        GRAPH.add((act, CM.isa, CM.Resume))
        GRAPH.add((act, CM.subject, agent))
        #GRAPH.add((act, CM.status, Literal("resume")))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Move))
        GRAPH.add((act1, CM.subject, agent))
        GRAPH.add((act1, CM.object, target))
    elif 'planned battle removed' in alert:
        GRAPH.add((act, CM.isa, CM.Cancel))
        GRAPH.add((act, CM.subject, agent))
        #GRAPH.add((act, CM.status, Literal("end")))
        if 'no real targets' in alert:
            GRAPH.add((act, CM.reason, Literal("no targets")))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Attack))
        GRAPH.add((act1, CM.subject, agent))
        GRAPH.add((act1, CM.object, target))
    elif 'firing has stopped'in alert:
        GRAPH.add((act, CM.isa, CM.Stop))
        #GRAPH.add((act, CM.subject, agent))
        #GRAPH.add((act, CM.Status, CM.END ))
        act1 = bnode('ACT')
        GRAPH.add((act, CM.object, act1 ))
        GRAPH.add((act1, CM.isa, CM.Attack))
        GRAPH.add((act1, CM.object, agent))
    elif 'detect' in alert:
        GRAPH.add((act, CM.isa, CM.Detect))
        GRAPH.add((act, CM.subject, agent))
        GRAPH.add((act, CM.subject, target))
        GRAPH.add((act, CM.reason, Literal(alert)))
    else:
        warn(f"Unrecognized alert: {alert}")
        unrecognized_alerts[alert] += 1
    return act

def create_initial_graph():
    global GRAPH, TASK, SIMULATION    
    # Create the inital RDF graph 
    GRAPH = Graph()
    #CM = Namespace('http://purl.org/artiamas/cm/')
    GRAPH.bind("cm", CM)
    GRAPH.bind("owl", OWL)
    GRAPH.bind("rdf", RDF)
    GRAPH.bind('rdfs', RDFS)
    SIMULATION = bnode('SIM')
    GRAPH.add((SIMULATION, RDF.type, CM.Simulation))
    TASK = bnode('TASK')
    GRAPH.add((TASK, RDF.type, CM.CrossingTask))
    GRAPH.add((SIMULATION, CM.task, TASK))
    return(GRAPH)

def json2graph (all_data):
    """ Builds and returns a graph representation of the messages """
    global GRAPH, TASK, SIMULATION
    GRAPH = create_initial_graph()
    messages = []
    actions = []
    for msg in all_data:
        subj = bnode('MSG')
        messages.append(subj)
        GRAPH.add((subj, RDF.type, CM.CombatMessage))
        #GRAPH.add((subj, CM.task, TASK))
        for prop, obj in msg.items():
            if prop not in str2property:
                warn(f"Unrecognized property {prop}")
                continue
            obj = str2object(msg, prop, obj)
            prop = str2property[prop]
            GRAPH.add((subj, prop, obj))
        # new properties
        GRAPH.add((subj, CM.minutes, Literal(time_to_minutes(msg['Time'])) ))
        act = add_action(msg, subj)
        actions.append(act)
    GRAPH.add((SIMULATION, CM.firstMessage, messages[0]))
    GRAPH.add((SIMULATION, CM.lastMessage, messages[-1]))
    GRAPH.add((SIMULATION, CM.firstAction, actions[0]))
    GRAPH.add((SIMULATION, CM.lastAction, actions[-1]))
    if link_msgs_actions:
        for i in range(len(messages)-1):
            GRAPH.add((messages[i], CM.nextMessage, messages[i+1]))
            GRAPH.add((actions[i], CM.nextAction, actions[i+1]))
    print(f"Added {len(messages)} messages and {len(actions)} actions")
    # we use CM.isa for immediate types, add rdf:type assersions
    for row in GRAPH.query("select ?X ?T {?X cm:isa ?T}"):
        GRAPH.add((row.X, RDF.type, row.T))
    return GRAPH

def sheet2graph(sheet_file):
    """ combine the simulation triples and the ontology triples, infer inherited type relations, and returns graph """
    global GRAPH
    GRAPH = json2graph(sheet2json(sheet_file))
    # read cm ontology into another graph from a local file or web
    if exists('cm.ttl'):
        gcm = Graph().parse("cm.ttl")
    else:
        gcm = Graph().parse("http://purl.org/artiamas/cm", format='ttl')
    g2 = GRAPH + gcm
    # Add inferred types
    for row in g2.query("select ?X ?ST {?X rdf:type/rdfs:subClassOf* ?ST}"):
        g2.add((row.X, RDF.type, row.ST))
    return(g2)
    #for key, value in unrecognized_alerts.items(): print(value, ':', key)    
